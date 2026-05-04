"""
Automated Weekly Sales Reporter

Processes daily Excel workbooks in the 'output/' directory to generate 
a 'week_summary' sheet featuring KPIs, item totals, and a comparative 
daily layout.
"""

import openpyxl, os
from openpyxl.styles import Font, Alignment
from pathlib import Path
from openpyxl.utils import get_column_letter


base_path = Path(__file__).parent

# Now we define folders relative to that base path
folder = base_path / "data"

output_folder = base_path / "output"

os.makedirs(output_folder, exist_ok=True)

bold_font = Font(bold=True)
center_align = Alignment(horizontal='center')


# sort the listing of files with sorted method
for file in sorted(os.listdir(folder)):

    # this conditional will skip any file that is not excel file
    if not file.endswith(".xlsx"):
        continue

    # the month, start_date, and end_date are needed later in the workbook
    month = file.split(".")[0].split("_")[0].title()
    start_date = file.split(".")[0].split("_")[1]
    end_date = file.split(".")[0].split("_")[3]

    week = f"{start_date}-{end_date}"

    # add data_only to return the actual value of the formula in any cell
    wb = openpyxl.load_workbook(folder/file, data_only=True)

    # to find the total sales for each day across all sheets
    for sheet in wb.sheetnames:
        daily_sheet = wb[sheet]

        max_row = daily_sheet.max_row
        max_col = daily_sheet.max_column

        daily_total = 0

        for row in range(2, max_row + 1):
            for col in range(max_col, max_col + 1):
    
            # conditional to check for and skip empty cells
            # in the total_price column
                if daily_sheet.cell(row, col).value is None:
                    continue

                daily_total += daily_sheet.cell(row, col).value

            daily_sheet.cell(max_row + 1, 1).font = bold_font
            daily_sheet.cell(max_row + 1, 1).value = "Daily Total"
            daily_sheet.cell(max_row + 1, max_col).font = bold_font
            daily_sheet.cell(max_row + 1, max_col).value = daily_total

    
    # dictionary for storing the daily total, to be later used for 
    # computing the best sales day
    totals_week = {}

    # initiate the weekly total at zero
    weekly_total = 0
    tot_week = 0

    # to find the total sales in a week
    for sheet in wb.sheetnames:
        daily_sheet = wb[sheet]

        max_row = daily_sheet.max_row
        max_col = daily_sheet.max_column

        # to get the cell for the daily_total calculated earlier
        for row in range(max_row, max_row + 1):
            for col in range(max_col, max_col+1):

                if daily_sheet.cell(row, col).value is None:
                    continue

                tot_week = daily_sheet.cell(row, col).value

        # add all the daily totals in a workbook and save them to weekly_total
        weekly_total += tot_week

        # adding the daily totals to a dictionary which will be 
        # used in finding the best performing day later
        if daily_sheet.title not in totals_week:
            totals_week[daily_sheet.title] = tot_week

    all_days_data = {}
    layout = {}
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    sheetnames = []

    # this dictionary will be used for gettting the total of each item
    # sold for each day then their total for the week
    week_items = {}
    
    # to find the total of each item sold for each day
    # we're calculating this here when it should have been the last 
    # thing to calculate. The reason for computing here is that we're 
    # calculating daily total and appending it to the row after the last
    # item row
    for sheet in wb.sheetnames:
        daily_sheet = wb[sheet]

        max_row = daily_sheet.max_row
        max_col = daily_sheet.max_column

        day_items = {}

        # to find the corresponding quantity sold for each item
        for row in range(2, max_row + 1):

            # Getting the items from each row
            for col in range(2, 3):
                if daily_sheet.cell(row, col).value is None:
                    continue

                item_name = daily_sheet.cell(row, col).value

            # Getting the corresponding quantity for each item
            for col in range(3, 4):
                if daily_sheet.cell(row, col).value is None:
                    continue

                quantity = daily_sheet.cell(row, col).value

                if item_name not in day_items:
                    day_items[item_name] = []
 
            # append the corresponding quantity of each item_name inside
            # items dictionary to the list of the item_name
            day_items[item_name].append(quantity)

        daily_sheet.cell(max_row + 3, 2).font = bold_font
        daily_sheet.cell(max_row + 3, 2).value = f"Summary of Sales for {daily_sheet.title}"

        daily_sheet.cell(max_row + 4, 2).font = bold_font
        daily_sheet.cell(max_row + 4, 2).value = "Item"
        daily_sheet.cell(max_row + 4, 3).font = bold_font
        daily_sheet.cell(max_row + 4, 3).value = "Total Quantity Sold Today"

        # initialise the starting point of filling the total daily
        # sales for each item
        current_row = max_row + 5

        for k, v in day_items.items():
            day_final = sum(v)

            daily_sheet.cell(current_row, 2).value = k
            daily_sheet.cell(current_row, 3).value = day_final

            current_row += 1

            if k not in week_items:
                week_items[k] = []

            week_items[k].append(day_final)

        all_days_data[daily_sheet.title] = day_items

        if daily_sheet.title not in sheetnames:
            sheetnames.append(daily_sheet.title)  
    
    for i, d in enumerate(days):
        if i < len(sheetnames):
            layout[d] = sheetnames[i]
        else:
            break

    wb.create_sheet("week_summary", index=0)
    week_summary = wb["week_summary"]

    week_summary.column_dimensions['a'].width = 20

    # Adjusting the width of the other columsn for better visuals
    for col in range(2, 2 + len(layout.keys())):
        col_letter = get_column_letter(col)
        week_summary.column_dimensions[col_letter].width = 12

    week_summary['a1'].font = bold_font
    week_summary['a1'] = "WEEKLY REPORT"
    week_summary.merge_cells('a1:b1')
    week_summary['a1'].alignment = center_align

    week_summary['a2'].font = bold_font
    week_summary['a2'] = "Period:"
    week_summary['b2'] = f"{month} {start_date} to {end_date}"

    week_summary['a3'].font = bold_font
    week_summary['a3'] = "Month:"
    week_summary['b3'] = month

    week_summary['a6'].font = bold_font
    week_summary['a6'] = "KPIs"
    week_summary.merge_cells('a6:b6')
    week_summary['a6'].alignment = center_align

    week_summary['a7'].font = bold_font
    week_summary['a7'] = "Total Revenue (NGN)"
    week_summary['b7'].font = bold_font
    week_summary['b7'] = weekly_total

    # finding the best perfoming day across all week
    best_day = max(totals_week, key=totals_week.get)

    # Finding the corresponding day for the best_day
    for y, z in layout.items():
        if best_day == z:
            actual_best_day = y
   
    week_summary['a8'].font = bold_font
    week_summary['a8'] = "Best Day:"
    week_summary['b8'] = actual_best_day 

    week_summary['a11'].font = bold_font
    week_summary['a11'] = "ITEM TOTALS"
    week_summary.merge_cells('a11:b11')
    week_summary['a11'].alignment = center_align

    week_summary['a12'] = "Item Name"
    week_summary['a12'] = "Qty Sold"
    
    
    current_row = 12
    for m, n in week_items.items():
        week_final = sum(n)

        week_summary.cell(current_row, 1).value = m        
        week_summary.cell(current_row, 2).value = week_final

        current_row += 1

    current_row += 2

    week_summary.cell(current_row, 1).font = bold_font
    week_summary.cell(current_row, 1).value = "COMPARATIVE LAYOUT"
    comp_layout_first_cell = week_summary.cell(current_row, 1).coordinate
    comp_layout_second_cell = week_summary.cell(current_row, 1 + len(layout.keys())).coordinate
    week_summary.merge_cells(f"{comp_layout_first_cell}:{comp_layout_second_cell}")
    week_summary[comp_layout_first_cell].alignment = center_align
    
    week_summary.cell(current_row + 1, 1).font = bold_font
    week_summary.cell(current_row + 1, 1).value = "Item Name"

    col = 2
    # We're using the exact number of days represented in the layout 
    # dictionary here so we don't have to change thing by hand
    for day in layout.keys():
        week_summary.cell(current_row + 1, col).font = bold_font
        week_summary.cell(current_row + 1, col).value = day

        col += 1

    current_row += 2

    # store the current row to where the items' daily total values will be printed
    grid_start_row = current_row

    # filling the item names on each row (later we'll come 
    # back to sort the order of this in ascending order)
    for items, quantities in week_items.items():
        week_summary.cell(current_row , 1).value = items

        current_row += 1

    # filling the grid for comparative layout
    for index, name in enumerate(week_items.keys()):
        row = grid_start_row + index
        col = 2

        for day in layout.keys():
            actual_sheet = layout[day]
            sheet_data = all_days_data.get(actual_sheet, {})

            # Checkk if item name exists in the sheet data
            if name in sheet_data:
                val = sum(sheet_data[name])
            else:
                val = 0

            week_summary.cell(row, col).value = val

            col += 1

    
    current_row += 2   

    week_summary.cell(current_row, 1).font = bold_font
    week_summary.cell(current_row, 1).value = "FINANCIAL SUMMARY"
    fin_sum_first_cell = week_summary.cell(current_row, 1).coordinate
    fin_sum_second_cell = week_summary.cell(current_row, 2).coordinate
    week_summary.merge_cells(f"{fin_sum_first_cell}:{fin_sum_second_cell}")
    week_summary[fin_sum_first_cell].alignment = center_align

    week_summary.cell(current_row + 1, 1).font = bold_font
    week_summary.cell(current_row + 1, 1).value = "Metric"
    week_summary.cell(current_row + 1, 2).font = bold_font
    week_summary.cell(current_row + 1, 2).value = "Value"

    week_summary.cell(current_row + 2, 1).value = "Average Daily Sale"
    week_summary.cell(current_row + 2, 2).value = round(weekly_total/len(sheetnames), 2)

    week_summary.cell(current_row + 3, 1).value = "Best Performing Day"
    week_summary.cell(current_row + 3, 2).value = actual_best_day 

    # empty dictionary to track the sales in a week which
    # which will be used to later compute the best selling item
    most_sold = {}

    for items, quantities in week_items.items():
        week_final = sum(quantities)

        if items not in most_sold:
            most_sold[items] = week_final

    most_sold_item = max(most_sold, key=most_sold.get)

    week_summary.cell(current_row + 4, 1).value = "Most Sold Item"
    week_summary.cell(current_row + 4, 2).value = most_sold_item


    print(f"Successfully processed the sales for week of {month} {start_date} to {end_date}")

    modified_file = os.path.join(output_folder, f"modified_{file}")

    wb.save(modified_file)
